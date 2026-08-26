from __future__ import annotations

import logging
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import to_month_label, write_json


plt.switch_backend("Agg")

CHART_INK = "#0b1f2a"
CHART_MUTED = "#6b7a84"
CHART_LINE = "#d7dbe0"
CHART_BAR = "#0f4c5c"
CHART_BAR_ALT = "#b45309"
CHART_ACCENT = "#0f4c5c"
CHART_BG = "#ffffff"

plt.rcParams.update({
    "font.family": ["Inter", "Helvetica Neue", "Helvetica", "Arial", "DejaVu Sans"],
    "font.size": 11,
    "axes.edgecolor": CHART_LINE,
    "axes.labelcolor": CHART_MUTED,
    "axes.titlecolor": CHART_INK,
    "axes.titleweight": "semibold",
    "axes.titlesize": 15,
    "axes.titlelocation": "left",
    "axes.titlepad": 14,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.spines.left": False,
    "axes.grid": True,
    "grid.color": CHART_LINE,
    "grid.linestyle": "-",
    "grid.linewidth": 0.6,
    "grid.alpha": 0.7,
    "xtick.color": CHART_MUTED,
    "ytick.color": CHART_MUTED,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "figure.facecolor": CHART_BG,
    "axes.facecolor": CHART_BG,
    "savefig.facecolor": CHART_BG,
    "savefig.bbox": "tight",
    "savefig.pad_inches": 0.25,
})


def _format_number_short(value: float, _pos=None) -> str:
    if value is None:
        return ""
    absval = abs(value)
    for divisor, suffix in ((1_000_000_000, "B"), (1_000_000, "M"), (1_000, "k")):
        if absval >= divisor:
            scaled = value / divisor
            text = f"{scaled:.1f}"
            if text.endswith(".0"):
                text = text[:-2]
            return f"{text}{suffix}"
    if value == int(value):
        return f"{int(value):,}"
    return f"{value:,.1f}"


def _format_full_number(value: float) -> str:
    if value is None:
        return ""
    if value == int(value):
        return f"{int(value):,}"
    return f"{value:,.1f}"


def write_outputs(
    output_dir: Path,
    data_quality_dir: Path,
    charts_dir: Path,
    website_dir: Path,
    flows: pd.DataFrame,
    country_summary: pd.DataFrame,
    routes: pd.DataFrame,
    transport_summary: pd.DataFrame,
    company_activity: pd.DataFrame,
    company_commodity: pd.DataFrame,
    company_market_context: pd.DataFrame,
    malformed_imports: pd.DataFrame,
    quality: dict,
) -> None:
    logging.info("Writing CSV outputs...")
    flows.to_csv(output_dir / "steel_import_flows.csv", index=False)
    country_summary.to_csv(output_dir / "country_summary.csv", index=False)
    routes.to_csv(output_dir / "origin_dispatch_routes.csv", index=False)
    transport_summary.to_csv(output_dir / "transport_summary.csv", index=False)
    company_activity.to_csv(output_dir / "company_activity.csv", index=False)
    company_commodity.to_csv(output_dir / "company_commodity_activity.csv", index=False)
    company_market_context.to_csv(output_dir / "company_market_context.csv", index=False)
    malformed_imports.to_csv(data_quality_dir / "malformed_import_records.csv", index=False)
    write_quality_report(output_dir / "data_quality_report.txt", quality)
    write_limitations(output_dir / "data_limitations.md")
    chart_paths = create_charts(charts_dir, flows, country_summary, transport_summary, company_activity, company_commodity, routes)
    build_workbook(
        output_dir / "UK_Steel_Import_Analysis.xlsx",
        country_summary,
        flows,
        routes,
        transport_summary,
        company_activity,
        company_commodity,
        quality,
        chart_paths,
    )
    write_website_summary(website_dir / "analysis-summary.json", flows, country_summary, transport_summary, company_activity, routes)


def write_quality_report(path: Path, quality: dict) -> None:
    lines = [f"{key}: {value}" for key, value in quality.items()]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_limitations(path: Path) -> None:
    content = """# Data limitations

Available from the public HMRC files:
- month
- commodity code and product description
- origin country
- dispatch country
- reported or inferred transport mode
- UK port/location code where present
- UK market tonnes and value
- importer monthly commodity activity

Not available from these files:
- exact supplier or steel mill
- purchase order number
- invoice-level shipment detail
- exact company-to-country linkage
- company shipment tonnage or value
- order date, departure date, or arrival date
- transit duration or Incoterms

Company-country and company-transport links are intentionally not fabricated in this analysis.
"""
    path.write_text(content, encoding="utf-8")


def create_charts(
    charts_dir: Path,
    flows: pd.DataFrame,
    country_summary: pd.DataFrame,
    transport_summary: pd.DataFrame,
    company_activity: pd.DataFrame,
    company_commodity: pd.DataFrame,
    routes: pd.DataFrame,
) -> dict[str, Path]:
    chart_paths: dict[str, Path] = {}
    chart_paths["top_countries_tonnes"] = bar_chart(
        charts_dir / "top_countries_tonnes.png",
        country_summary.head(20),
        "origin_country",
        "total_tonnes",
        "Top 20 origin countries by tonnes",
        color=CHART_BAR,
        value_suffix=" t",
    )
    chart_paths["top_countries_value"] = bar_chart(
        charts_dir / "top_countries_value.png",
        country_summary.head(20),
        "origin_country",
        "total_value",
        "Top 20 origin countries by value",
        color=CHART_BAR_ALT,
        value_prefix="£",
    )
    products = (
        flows.groupby("steel_product_group")["tonnes"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )
    chart_paths["product_groups"] = bar_chart(
        charts_dir / "top_product_groups.png",
        products.head(12),
        "steel_product_group",
        "tonnes",
        "Top steel product groups by tonnes",
        color=CHART_BAR,
        value_suffix=" t",
    )
    transport = (
        transport_summary.groupby("estimated_transport_mode")["tonnes"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )
    chart_paths["transport_modes"] = bar_chart(
        charts_dir / "transport_modes.png",
        transport,
        "estimated_transport_mode",
        "tonnes",
        "Steel tonnes by transport mode",
        color=CHART_BAR,
        value_suffix=" t",
    )
    monthly_importers = (
        company_commodity.groupby("first_seen").size().reset_index(name="new_company_commodity_links")
        if not company_commodity.empty
        else pd.DataFrame(columns=["first_seen", "new_company_commodity_links"])
    )
    activity_by_month = (
        company_activity.assign(period_label=company_activity["last_seen"].map(to_month_label))
        .groupby("period_label")
        .size()
        .reset_index(name="active_companies")
        .sort_values("period_label")
    )
    chart_paths["active_importers"] = line_chart(
        charts_dir / "active_importers_by_month.png",
        activity_by_month,
        "period_label",
        "active_companies",
        "Active steel importers by latest-seen month",
        color=CHART_ACCENT,
    )
    route_frame = (
        routes.assign(route=routes["origin_country_name"] + " -> " + routes["dispatch_country_name"])
        .head(15)
    )
    chart_paths["routes"] = bar_chart(
        charts_dir / "top_routes.png",
        route_frame,
        "route",
        "tonnes",
        "Top origin to dispatch routes",
        color=CHART_INK,
        value_suffix=" t",
    )
    return chart_paths


def bar_chart(
    path: Path,
    frame: pd.DataFrame,
    category_col: str,
    value_col: str,
    title: str,
    color: str = CHART_BAR,
    value_prefix: str = "",
    value_suffix: str = "",
) -> Path:
    categories = frame[category_col].astype(str).tolist()
    values = frame[value_col].fillna(0).astype(float).tolist()

    n = max(len(categories), 1)
    height = max(5.5, 0.42 * n + 2.2)
    fig, ax = plt.subplots(figsize=(12, height))

    bars = ax.barh(categories, values, color=color, height=0.72, edgecolor="none")
    ax.set_title(title)
    ax.invert_yaxis()

    ax.xaxis.set_major_formatter(FuncFormatter(_format_number_short))
    ax.tick_params(axis="x", which="both", length=0, pad=6)
    ax.tick_params(axis="y", which="both", length=0, pad=8)

    ax.grid(axis="x", color=CHART_LINE, linewidth=0.6, alpha=0.9)
    ax.grid(axis="y", visible=False)
    ax.set_axisbelow(True)

    for spine in ax.spines.values():
        spine.set_visible(False)

    max_val = max(values) if values else 0
    offset = max_val * 0.01 if max_val else 0
    for bar, value in zip(bars, values):
        label = f"{value_prefix}{_format_full_number(value)}{value_suffix}"
        ax.text(
            bar.get_width() + offset,
            bar.get_y() + bar.get_height() / 2,
            label,
            va="center",
            ha="left",
            fontsize=9.5,
            color=CHART_INK,
        )

    if max_val:
        ax.set_xlim(0, max_val * 1.14)

    for label in ax.get_yticklabels():
        label.set_color(CHART_INK)
        label.set_fontweight("medium")

    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def line_chart(
    path: Path,
    frame: pd.DataFrame,
    x_col: str,
    y_col: str,
    title: str,
    color: str = CHART_ACCENT,
) -> Path:
    x = frame[x_col].astype(str).tolist()
    y = frame[y_col].fillna(0).astype(float).tolist()

    fig, ax = plt.subplots(figsize=(12, 5.5))
    ax.plot(x, y, color=color, linewidth=2.2, marker="o", markersize=4.5, markerfacecolor=color, markeredgecolor=CHART_BG, markeredgewidth=1.2)
    ax.fill_between(range(len(x)), y, color=color, alpha=0.08)

    ax.set_title(title)

    ax.yaxis.set_major_formatter(FuncFormatter(_format_number_short))
    ax.tick_params(axis="x", which="both", length=0, pad=6)
    ax.tick_params(axis="y", which="both", length=0, pad=6)

    ax.grid(axis="y", color=CHART_LINE, linewidth=0.6, alpha=0.9)
    ax.grid(axis="x", visible=False)
    ax.set_axisbelow(True)

    for spine in ax.spines.values():
        spine.set_visible(False)

    step = max(1, len(x) // 12)
    ax.set_xticks(range(0, len(x), step))
    ax.set_xticklabels([x[i] for i in range(0, len(x), step)], rotation=35, ha="right")

    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def build_workbook(
    path: Path,
    country_summary: pd.DataFrame,
    flows: pd.DataFrame,
    routes: pd.DataFrame,
    transport_summary: pd.DataFrame,
    company_activity: pd.DataFrame,
    company_commodity: pd.DataFrame,
    quality: dict,
    chart_paths: dict[str, Path],
) -> None:
    workbook = Workbook()
    sheets = {
        "Overview": pd.DataFrame(
            [
                {"metric": "Total steel tonnes", "value": flows["tonnes"].sum()},
                {"metric": "Total steel import value", "value": flows["statistical_value_gbp"].sum()},
                {"metric": "Origin countries", "value": flows["origin_country_name"].nunique()},
                {"metric": "Active steel importers", "value": company_activity["company_id"].nunique()},
                {"metric": "Reliable transport share", "value": flows["transport_mode_reliable"].mean()},
            ]
        ),
        "Countries": country_summary,
        "Products": (
            flows.groupby(["commodity_code", "commodity_description", "steel_product_group"], dropna=False)
            .agg(tonnes=("tonnes", "sum"), value=("statistical_value_gbp", "sum"))
            .reset_index()
            .sort_values("tonnes", ascending=False)
        ),
        "Origin-Dispatch": routes,
        "Transport": transport_summary,
        "Importers": company_activity,
        "Company Activity": company_activity,
        "Company Products": company_commodity,
        "Data Quality": pd.DataFrame([quality]),
        "Methodology": pd.DataFrame(
            [
                {"note": "Origin and dispatch are kept separate throughout the analysis."},
                {"note": "Only Chapter 72 commodity codes are included in this version."},
                {"note": "Company-country shipment links are not inferred from public HMRC files."},
                {"note": "2026 is treated as year-to-date and should be compared with equivalent 2025 months."},
            ]
        ),
    }

    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for name, frame in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        write_dataframe_sheet(worksheet, frame)

    overview = workbook["Overview"]
    anchor_positions = {
        "top_countries_tonnes": "D2",
        "product_groups": "D22",
        "transport_modes": "N2",
        "routes": "N22",
    }
    for chart_name, cell in anchor_positions.items():
        if chart_name in chart_paths:
            overview.add_image(XLImage(str(chart_paths[chart_name])), cell)

    workbook.save(path)


def write_dataframe_sheet(worksheet, frame: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)

    if frame.empty:
        worksheet.append(["No data"])
        return

    worksheet.append(list(frame.columns))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in frame.itertuples(index=False, name=None):
        worksheet.append(list(row))

    worksheet.auto_filter.ref = worksheet.dimensions
    for idx, column in enumerate(frame.columns, start=1):
        max_len = max(len(str(column)), *(len(str(value)) for value in frame[column].head(200)))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 40)


def write_website_summary(
    path: Path,
    flows: pd.DataFrame,
    country_summary: pd.DataFrame,
    transport_summary: pd.DataFrame,
    company_activity: pd.DataFrame,
    routes: pd.DataFrame,
) -> None:
    top_products = (
        flows.groupby(["steel_product_group"], dropna=False)["tonnes"]
        .sum()
        .sort_values(ascending=False)
        .head(8)
        .reset_index()
        .to_dict(orient="records")
    )
    top_routes = (
        routes.head(10)
        .assign(route=lambda df: df["origin_country_name"] + " -> " + df["dispatch_country_name"])[
            ["route", "commodity_code", "steel_product_group", "tonnes", "value"]
        ]
        .to_dict(orient="records")
    )
    top_countries = country_summary.head(12)[
        ["origin_country", "total_tonnes", "total_value", "share_of_uk_steel_import_tonnes", "main_product_group"]
    ].to_dict(orient="records")
    transport_mix = (
        transport_summary.groupby("estimated_transport_mode")["tonnes"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
        .to_dict(orient="records")
    )
    status_mix = (
        company_activity.groupby("company_status")["company_id"]
        .nunique()
        .sort_values(ascending=False)
        .reset_index(name="companies")
        .to_dict(orient="records")
    )
    payload = {
        "generatedAt": pd.Timestamp.utcnow().isoformat(),
        "coverage": {
            "firstPeriod": flows["period"].min(),
            "lastPeriod": flows["period"].max(),
            "monthsCovered": int(flows["period"].nunique()),
        },
        "kpis": {
            "totalTonnes": round(float(flows["tonnes"].sum()), 2),
            "totalValueGbp": round(float(flows["statistical_value_gbp"].sum()), 2),
            "originCountries": int(flows["origin_country_name"].nunique()),
            "activeImporters": int(company_activity["company_id"].nunique()),
            "reliableTransportShare": round(float(flows["transport_mode_reliable"].mean()), 3),
        },
        "topCountries": top_countries,
        "topProducts": top_products,
        "transportMix": transport_mix,
        "companyStatusMix": status_mix,
        "topRoutes": top_routes,
    }
    write_json(path, payload)
